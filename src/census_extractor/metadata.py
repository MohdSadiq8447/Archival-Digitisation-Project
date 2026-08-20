"""Authoritative workbook metadata and provenance lookup."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class MetadataError(ValueError):
    """Raised when workbook provenance is missing or ambiguous."""


@dataclass(frozen=True, slots=True)
class DocumentMetadata:
    sheet: str
    pdf_id: str
    file_name: str
    state: str
    district: str
    year: int
    format_id: str
    source_page_start: int
    source_page_end: int
    table_id: str
    anchor_page: int
    continuation_page: int
    anchor_printed_page: int | None
    continuation_printed_page: int | None
    workbook_path: str
    workbook_sha256: str

    def provenance_dict(self) -> dict[str, Any]:
        return asdict(self)


def _optional_int(value: Any) -> int | None:
    return None if value is None or value == "" else int(value)


def _required_int(value: object, field_name: str) -> int:
    if value is None or value == "":
        raise MetadataError(f"Workbook field {field_name!r} is required")
    if not isinstance(value, (str, int, float)):
        raise MetadataError(f"Workbook field {field_name!r} is not scalar")
    return int(value)


class MetadataRegistry:
    REQUIRED_COLUMNS = {
        "pdf_id",
        "file_name",
        "state",
        "district",
        "year",
        "format_id",
        "source_page_start",
        "source_page_end",
        "table_id",
        "anchor_page",
        "continuation_page",
        "anchor_printed_page",
        "continuation_printed_page",
    }

    def __init__(self, workbook_path: Path):
        import hashlib

        self.workbook_path = Path(workbook_path).resolve()
        if not self.workbook_path.is_file():
            raise FileNotFoundError(f"Metadata workbook not found: {self.workbook_path}")
        self.workbook_sha256 = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        self._by_file: dict[str, DocumentMetadata] = {}
        self._by_id: dict[str, DocumentMetadata] = {}
        self._load()

    def _load(self) -> None:
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if headers is None:
                    continue
                names = [str(value).strip() if value is not None else "" for value in headers]
                missing = self.REQUIRED_COLUMNS.difference(names)
                if missing:
                    raise MetadataError(
                        f"Sheet {sheet.title!r} is missing columns: {sorted(missing)}"
                    )
                for values in rows:
                    raw = dict(zip(names, values, strict=True))
                    if not raw.get("file_name"):
                        continue
                    record = DocumentMetadata(
                        sheet=sheet.title,
                        pdf_id=str(raw["pdf_id"]).strip(),
                        file_name=str(raw["file_name"]).strip(),
                        state=str(raw["state"]).strip(),
                        district=str(raw["district"]).strip(),
                        year=_required_int(raw["year"], "year"),
                        format_id=str(raw["format_id"]).strip(),
                        source_page_start=_required_int(
                            raw["source_page_start"], "source_page_start"
                        ),
                        source_page_end=_required_int(raw["source_page_end"], "source_page_end"),
                        table_id=str(raw["table_id"]).strip(),
                        anchor_page=_required_int(raw["anchor_page"], "anchor_page"),
                        continuation_page=_required_int(
                            raw["continuation_page"], "continuation_page"
                        ),
                        anchor_printed_page=_optional_int(raw["anchor_printed_page"]),
                        continuation_printed_page=_optional_int(raw["continuation_printed_page"]),
                        workbook_path=str(self.workbook_path),
                        workbook_sha256=self.workbook_sha256,
                    )
                    file_key, id_key = record.file_name.casefold(), record.pdf_id.casefold()
                    if file_key in self._by_file or id_key in self._by_id:
                        raise MetadataError(f"Duplicate workbook record for {record.file_name!r}")
                    self._by_file[file_key] = record
                    self._by_id[id_key] = record
        finally:
            workbook.close()

    def get_for_pdf(self, pdf_path: Path) -> DocumentMetadata:
        record = self._by_file.get(Path(pdf_path).name.casefold())
        if record is None:
            raise MetadataError(f"No authoritative metadata row for PDF {Path(pdf_path).name!r}")
        return record

    def get(self, pdf_id: str) -> DocumentMetadata | None:
        return self._by_id.get(pdf_id.casefold())

    def all(self) -> list[DocumentMetadata]:
        return sorted(self._by_file.values(), key=lambda item: item.file_name.casefold())
