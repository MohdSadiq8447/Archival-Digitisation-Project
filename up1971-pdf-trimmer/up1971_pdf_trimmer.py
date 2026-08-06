from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

import pdf_inspector


PROJECT_ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = PROJECT_ROOT.parent
DEFAULT_WORKBOOK = WORKSPACE_ROOT / "UP_1971_Amenities_Page_Plan_With_Totals.xlsx"
DEFAULT_PDF_ROOT = WORKSPACE_ROOT / "1971-20260725T134344Z-1-001" / "1971" / "Uttar Pradesh"
DEFAULT_OUTPUT = PROJECT_ROOT / "python-output"

CATEGORY_DIRS = {
    "civic": "Civic Amenities",
    "mededu": "Medical_Education",
    "tehsil": "Tehsil_Appendix",
    "villages": "Villages",
}


@dataclass
class PageRange:
    start: int
    end: int


@dataclass
class VillageRange:
    tahsil: str
    printed_range: PageRange


@dataclass
class DistrictPlan:
    district: str
    source_pdf: str
    town_civic: PageRange | None = None
    town_mededu: PageRange | None = None
    tehsil_appendix: PageRange | None = None
    villages: list[VillageRange] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def normalize_district(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def slugify(value: str) -> str:
    text = value.strip().lower().replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def parse_range(value: Any) -> PageRange | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = re.fullmatch(r"(\d+)\s*-\s*(\d+)", text)
    if match:
        return PageRange(int(match.group(1)), int(match.group(2)))
    match = re.fullmatch(r"(\d+)", text)
    if match:
        number = int(match.group(1))
        return PageRange(number, number)
    return None


def load_trim_plan(workbook_path: Path) -> dict[str, DistrictPlan]:
    workbook = load_workbook(workbook_path, data_only=True)
    tehsil_sheet = workbook["Tehsil_and_Town"]
    village_sheet = workbook["Village_Level"]

    plan: dict[str, DistrictPlan] = {}

    tehsil_headers = [cell.value for cell in tehsil_sheet[1]]
    for row in tehsil_sheet.iter_rows(min_row=2, values_only=True):
        row_map = dict(zip(tehsil_headers, row))
        district = row_map.get("District")
        source_pdf = row_map.get("Source PDF")
        if not district or not source_pdf:
            continue

        key = normalize_district(district)
        spec = plan.get(key) or DistrictPlan(district=str(district), source_pdf=str(source_pdf))
        spec.town_civic = parse_range(row_map.get("Town Statement IV: Civic amenities pages"))
        spec.town_mededu = parse_range(row_map.get("Town Statement V: Medical/Education pages"))
        spec.tehsil_appendix = parse_range(row_map.get("Tahsil appendix pages"))
        note = row_map.get("Verification note")
        if note:
            spec.notes.append(str(note).strip())
        plan[key] = spec

    village_headers = [cell.value for cell in village_sheet[1]]
    for row in village_sheet.iter_rows(min_row=2, values_only=True):
        row_map = dict(zip(village_headers, row))
        district = row_map.get("District")
        source_pdf = row_map.get("Source PDF")
        tahsil = row_map.get("Tahsil")
        if not district or not source_pdf or not tahsil:
            continue
        if str(tahsil).strip().upper() == "DISTRICT TOTAL":
            continue

        key = normalize_district(district)
        spec = plan.get(key) or DistrictPlan(district=str(district), source_pdf=str(source_pdf))
        spec.villages.append(
            VillageRange(
                tahsil=str(tahsil),
                printed_range=PageRange(
                    int(row_map["Printed page start"]),
                    int(row_map["Printed page end"]),
                ),
            )
        )
        note = row_map.get("Verification note")
        if note:
            spec.notes.append(str(note).strip())
        plan[key] = spec

    for spec in plan.values():
        spec.notes = sorted(set(filter(None, spec.notes)))

    return plan


def build_page_texts(
    pdf_path: Path,
    target_pages: list[int] | None = None,
    use_region_fallback: bool = True,
) -> dict[int, str]:
    page_texts: dict[int, str] = {}
    reader = PdfReader(str(pdf_path))
    selected_pages = target_pages if target_pages is not None else list(range(len(reader.pages)))
    page_regions = []
    for page_index in selected_pages:
        page = reader.pages[page_index]
        text = page.extract_text() or ""
        if text.strip():
            page_texts[page_index] = text
            continue
        box = page.mediabox
        page_regions.append((page_index, [(0, 0, float(box.width), float(box.height))]))

    if not page_regions or not use_region_fallback:
        return page_texts

    chunk_size = 16
    for start in range(0, len(page_regions), chunk_size):
        chunk = page_regions[start : start + chunk_size]
        extracted = pdf_inspector.extract_text_in_regions(str(pdf_path), chunk)
        for page_region in extracted:
            if not page_region.regions:
                continue
            page_texts[int(page_region.page)] = page_region.regions[0].text or ""
    return page_texts


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().lower()


def leading_page_number(text: str) -> int | None:
    snippet = re.sub(r"\s+", " ", text[:120]).strip()
    match = re.match(r"^(\d{1,4})\b", snippet)
    return int(match.group(1)) if match else None


def find_marker(page_texts: dict[int, str], page_indexes: list[int], matcher) -> dict[str, Any] | None:
    for page_index in page_indexes:
        page_text = page_texts.get(page_index, "")
        if not page_text:
            continue
        if not matcher(normalize_text(page_text)):
            continue
        printed_page = leading_page_number(page_text)
        if printed_page is None:
            continue
        return {
            "actual_page": page_index + 1,
            "printed_page": printed_page,
            "preview": re.sub(r"\s+", " ", page_text[:180]).strip(),
        }
    return None


def collect_page_anchors(page_texts: dict[int, str], page_indexes: list[int]) -> list[dict[str, Any]]:
    anchors: list[dict[str, Any]] = []
    for page_index in page_indexes:
        page_text = page_texts.get(page_index, "")
        if not page_text:
            continue
        printed_page = leading_page_number(page_text)
        if printed_page is None or printed_page < 1 or printed_page > 1000:
            continue
        normalized = normalize_text(page_text)
        anchors.append(
            {
                "actual_page": page_index + 1,
                "printed_page": printed_page,
                "offset": page_index + 1 - printed_page,
                "is_town_civic": "town statement" in normalized and "civic" in normalized,
                "is_town_mededu": "town statement" in normalized and "medical" in normalized,
                "is_village": "village" in normalized and ("amenit" in normalized or "directory" in normalized),
                "is_appendix": "appendix" in normalized and "tahsil" in normalized,
                "preview": re.sub(r"\s+", " ", page_text[:180]).strip(),
            }
        )
    return anchors


def range_contains_printed(page_range: PageRange, printed_page: int, padding: int = 2) -> bool:
    return page_range.start - padding <= printed_page <= page_range.end + padding


def infer_offset(
    section: str,
    printed_ranges: list[PageRange],
    page_count: int,
    page_texts: dict[int, str],
    anchors: list[dict[str, Any]],
    title_matcher=None,
    preferred_offsets: list[int | None] | None = None,
) -> dict[str, Any]:
    preferred_offsets = preferred_offsets or []
    candidate_offsets = {anchor["offset"] for anchor in anchors}
    candidate_offsets.update(offset for offset in preferred_offsets if offset is not None)

    candidates: list[dict[str, Any]] = []
    if not printed_ranges:
        return {
            "section": section,
            "chosen_offset": None,
            "confidence": 0.0,
            "reason": "no_ranges",
            "candidates": [],
        }

    min_printed = min(page_range.start for page_range in printed_ranges)
    max_printed = max(page_range.end for page_range in printed_ranges)

    for offset in sorted(candidate_offsets):
        resolved_ranges = [resolve_range(page_range, offset) for page_range in printed_ranges]
        usable_ranges = [
            PageRange(max(1, page_range.start), min(page_count, page_range.end))
            for page_range in resolved_ranges
        ]
        valid_ranges = [
            page_range
            for page_range in resolved_ranges
            if page_range.start >= 1 and page_range.end >= page_range.start and page_range.end <= page_count
        ]
        overflow_pages = sum(
            max(0, 1 - page_range.start) + max(0, page_range.end - page_count)
            for page_range in resolved_ranges
        )
        small_end_overflow = (
            overflow_pages <= 2
            and all(page_range.start >= 1 and page_range.start <= page_count for page_range in resolved_ranges)
            and all(page_range.end >= page_range.start for page_range in usable_ranges)
        )
        nearby_anchors = [
            anchor
            for anchor in anchors
            if anchor["offset"] == offset and min_printed - 10 <= anchor["printed_page"] <= max_printed + 10
        ]
        range_anchors = [
            anchor
            for anchor in nearby_anchors
            if any(range_contains_printed(page_range, anchor["printed_page"]) for page_range in printed_ranges)
        ]
        section_anchors = [
            anchor
            for anchor in range_anchors
            if (
                section == "town"
                and (anchor["is_town_civic"] or anchor["is_town_mededu"])
                or section == "villages"
                and anchor["is_village"]
                or section == "tehsil"
                and (anchor["is_appendix"] or anchor["is_village"])
            )
        ]

        title_hits = 0
        if title_matcher is not None:
            for page_range in resolved_ranges:
                scan_start = max(1, page_range.start)
                scan_end = min(page_count, page_range.start + 3)
                for actual_page in range(scan_start, scan_end + 1):
                    if title_matcher(normalize_text(page_texts.get(actual_page - 1, ""))):
                        title_hits += 1
                        break

        continuity_hits = 0
        sorted_anchors = sorted(range_anchors, key=lambda anchor: anchor["actual_page"])
        for previous, current in zip(sorted_anchors, sorted_anchors[1:]):
            if current["actual_page"] > previous["actual_page"] and current["printed_page"] >= previous["printed_page"]:
                continuity_hits += 1

        all_valid = len(valid_ranges) == len(resolved_ranges)
        usable = all_valid or small_end_overflow
        preferred_bonus = 1 if offset in preferred_offsets else 0
        score = (
            len(valid_ranges) * 1000
            + (500 if all_valid else 0)
            + (250 if usable and not all_valid else 0)
            + len(range_anchors) * 25
            + len(section_anchors) * 20
            + title_hits * 15
            + continuity_hits * 5
            + preferred_bonus * 10
            - overflow_pages * 20
        )
        candidates.append(
            {
                "offset": offset,
                "score": score,
                "valid_ranges": len(valid_ranges),
                "total_ranges": len(resolved_ranges),
                "all_ranges_valid": all_valid,
                "usable": usable,
                "small_end_overflow": small_end_overflow and not all_valid,
                "range_anchor_count": len(range_anchors),
                "section_anchor_count": len(section_anchors),
                "title_hits": title_hits,
                "continuity_hits": continuity_hits,
                "overflow_pages": overflow_pages,
                "resolved_ranges": [page_range.__dict__ for page_range in resolved_ranges],
                "usable_ranges": [page_range.__dict__ for page_range in usable_ranges],
                "sample_anchors": range_anchors[:5],
            }
        )

    candidates.sort(
        key=lambda candidate: (
            candidate["usable"],
            candidate["score"],
            candidate["section_anchor_count"],
            candidate["range_anchor_count"],
            -abs(candidate["offset"]),
        ),
        reverse=True,
    )
    chosen = candidates[0] if candidates else None
    if not chosen:
        return {
            "section": section,
            "chosen_offset": None,
            "confidence": 0.0,
            "reason": "no_candidate_offsets",
            "candidates": [],
        }

    confidence = (
        0.95
        if chosen["all_ranges_valid"] and chosen["range_anchor_count"] >= 3
        else 0.85
        if chosen["small_end_overflow"] and chosen["range_anchor_count"] >= 3
        else 0.75
        if chosen["usable"]
        else 0.35
    )
    return {
        "section": section,
        "chosen_offset": chosen["offset"],
        "confidence": confidence,
        "reason": "ok"
        if chosen["all_ranges_valid"]
        else "minor_pdf_end_overflow_clamped"
        if chosen["small_end_overflow"]
        else "best_candidate_has_out_of_bounds_ranges",
        "chosen": chosen,
        "candidates": candidates[:8],
    }


def resolve_range(page_range: PageRange, offset: int) -> PageRange:
    return PageRange(page_range.start + offset, page_range.end + offset)


def expand_ranges(ranges: list[PageRange]) -> list[int]:
    pages: list[int] = []
    for page_range in ranges:
        pages.extend(range(page_range.start, page_range.end + 1))
    return pages


def pages_around_ranges(ranges: list[PageRange], offset: int | None, page_count: int, padding: int = 3) -> list[int]:
    if offset is None:
        return []
    pages: list[int] = []
    for page_range in ranges:
        for printed_page in (page_range.start, page_range.end):
            actual_page = printed_page + offset
            for page_number in range(actual_page - padding, actual_page + padding + 1):
                if 1 <= page_number <= page_count:
                    pages.append(page_number - 1)
    return pages


def ranges_from_inference(inference: dict[str, Any], printed_ranges: list[PageRange]) -> list[PageRange] | None:
    chosen = inference.get("chosen")
    if not chosen or not chosen.get("usable"):
        return None
    range_dicts = chosen.get("usable_ranges") or chosen.get("resolved_ranges") or []
    if len(range_dicts) != len(printed_ranges):
        return None
    return [PageRange(int(item["start"]), int(item["end"])) for item in range_dicts]


def write_trimmed_pdf(source_pdf: Path, ranges: list[PageRange], output_pdf: Path) -> None:
    reader = PdfReader(str(source_pdf))
    writer = PdfWriter()
    for page_range in ranges:
        for page_number in range(page_range.start, page_range.end + 1):
            writer.add_page(reader.pages[page_number - 1])
    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    with output_pdf.open("wb") as handle:
        writer.write(handle)


def category_filename(district: str, category: str) -> str:
    slug = slugify(district)
    if category == "civic":
        return f"{slug}_civic_1971.pdf"
    if category == "mededu":
        return f"{slug}_mededu_1971.pdf"
    if category == "tehsil":
        return f"{slug}_tehsil_1971.pdf"
    return f"{slug}_villages_1971.pdf"


def process_district(plan: DistrictPlan, pdf_root: Path, output_root: Path, sheet_filter: str | None) -> dict[str, Any]:
    pdf_path = pdf_root / plan.source_pdf
    classification = pdf_inspector.classify_pdf(str(pdf_path))
    page_count = int(classification.page_count)

    early_pages = list(range(min(page_count, 110)))
    late_pages = list(range(max(0, page_count - 50), page_count))
    page_texts = build_page_texts(pdf_path, sorted(set(early_pages + late_pages)), use_region_fallback=False)

    town_marker = find_marker(
        page_texts,
        early_pages,
        lambda text: "town statement" in text and "civic and other" in text,
    )
    med_marker = find_marker(
        page_texts,
        early_pages,
        lambda text: "town statement" in text and "medical" in text,
    )
    village_marker = find_marker(
        page_texts,
        early_pages,
        lambda text: "village" in text and "amenities and" in text,
    )
    appendix_marker = find_marker(
        page_texts,
        late_pages,
        lambda text: "appendix" in text and "tahsil" in text and "abstract" in text,
    )

    raw_town_offset = (
        town_marker["actual_page"] - town_marker["printed_page"]
        if town_marker
        else med_marker["actual_page"] - med_marker["printed_page"]
        if med_marker
        else None
    )
    raw_village_offset = village_marker["actual_page"] - village_marker["printed_page"] if village_marker else None
    raw_appendix_offset = appendix_marker["actual_page"] - appendix_marker["printed_page"] if appendix_marker else None

    town_ranges = [page_range for page_range in [plan.town_civic, plan.town_mededu] if page_range is not None]
    village_ranges_for_scan = [item.printed_range for item in plan.villages]
    focused_pages = sorted(
        set(early_pages + late_pages)
        | set(pages_around_ranges(town_ranges, raw_town_offset, page_count))
        | set(pages_around_ranges(village_ranges_for_scan, raw_village_offset, page_count))
        | set(pages_around_ranges([plan.tehsil_appendix] if plan.tehsil_appendix else [], raw_appendix_offset, page_count))
    )
    missing_focused_pages = [page for page in focused_pages if page not in page_texts]
    if missing_focused_pages:
        page_texts.update(build_page_texts(pdf_path, missing_focused_pages, use_region_fallback=False))
    page_anchors = collect_page_anchors(page_texts, focused_pages)

    town_inference = infer_offset(
        "town",
        town_ranges,
        page_count,
        page_texts,
        page_anchors,
        lambda text: "town statement" in text and ("civic" in text or "medical" in text),
        [raw_town_offset, raw_village_offset],
    )
    village_inference = infer_offset(
        "villages",
        [item.printed_range for item in plan.villages],
        page_count,
        page_texts,
        page_anchors,
        lambda text: "village" in text and ("amenit" in text or "directory" in text),
        [raw_village_offset, raw_town_offset],
    )
    appendix_inference = infer_offset(
        "tehsil",
        [plan.tehsil_appendix] if plan.tehsil_appendix is not None else [],
        page_count,
        page_texts,
        page_anchors,
        lambda text: "appendix" in text and "tahsil" in text,
        [raw_appendix_offset, village_inference["chosen_offset"], town_inference["chosen_offset"]],
    )

    town_offset = town_inference["chosen_offset"]
    village_offset = village_inference["chosen_offset"]
    appendix_offset = appendix_inference["chosen_offset"]

    report: dict[str, Any] = {
        "district": plan.district,
        "source_pdf": plan.source_pdf,
        "source_path": str(pdf_path),
        "page_count": page_count,
        "pdf_type": classification.pdf_type,
        "confidence": classification.confidence,
        "pages_needing_ocr": [page + 1 for page in classification.pages_needing_ocr],
        "markers": {
            "town_marker": town_marker,
            "med_marker": med_marker,
            "village_marker": village_marker,
            "appendix_marker": appendix_marker,
        },
        "detected_anchors": page_anchors[:250],
        "offset_inference": {
            "town": town_inference,
            "villages": village_inference,
            "tehsil": appendix_inference,
        },
        "notes": plan.notes,
        "warnings": [],
        "outputs": {},
    }

    if town_offset is None and sheet_filter != "Village_Level":
        report["warnings"].append("Could not infer town-section offset.")
    if village_offset is None and sheet_filter != "Tehsil_and_Town":
        report["warnings"].append("Could not infer village-section offset.")
    if appendix_marker is None and sheet_filter != "Village_Level":
        report["warnings"].append("Could not find appendix marker; using scored fallback offset.")
    for section_name, inference in report["offset_inference"].items():
        if inference["reason"] != "ok":
            report["warnings"].append(f"{section_name}: {inference['reason']}.")

    targets: list[tuple[str, list[PageRange], list[dict[str, Any]] | None]] = []
    if sheet_filter in (None, "Tehsil_and_Town"):
        town_resolved_ranges = ranges_from_inference(town_inference, town_ranges)
        if plan.town_civic and town_offset is not None and town_resolved_ranges:
            targets.append(("civic", [town_resolved_ranges[0]], None))
        if plan.town_mededu and town_offset is not None and town_resolved_ranges:
            town_mededu_index = 1 if plan.town_civic else 0
            targets.append(("mededu", [town_resolved_ranges[town_mededu_index]], None))
        appendix_resolved_ranges = ranges_from_inference(
            appendix_inference,
            [plan.tehsil_appendix] if plan.tehsil_appendix is not None else [],
        )
        if plan.tehsil_appendix and appendix_offset is not None and appendix_resolved_ranges:
            targets.append(("tehsil", appendix_resolved_ranges, None))

    if sheet_filter in (None, "Village_Level") and plan.villages and village_offset is not None:
        village_ranges = ranges_from_inference(village_inference, [item.printed_range for item in plan.villages])
        if village_ranges is None:
            village_ranges = [resolve_range(item.printed_range, village_offset) for item in plan.villages]
        tahsils = [
            {
                "tahsil": item.tahsil,
                "printed_range": item.printed_range.__dict__,
                "actual_range": village_ranges[index].__dict__,
            }
            for index, item in enumerate(plan.villages)
        ]
        targets.append(("villages", village_ranges, tahsils))

    pages_needing_ocr = set(classification.pages_needing_ocr)

    for category, ranges, tahsils in targets:
        if any(page_range.start < 1 or page_range.end < page_range.start or page_range.end > page_count for page_range in ranges):
            report["warnings"].append(
                f"{category}: resolved page range is outside the PDF bounds for {plan.district}."
            )
            report["outputs"][category] = {
                "status": "skipped",
                "resolved_ranges": [page_range.__dict__ for page_range in ranges],
                "reason": "out_of_bounds",
            }
            continue

        output_path = output_root / CATEGORY_DIRS[category] / category_filename(plan.district, category)
        write_trimmed_pdf(pdf_path, ranges, output_path)
        category_pages = expand_ranges(ranges)
        ocr_risk_pages = sorted(page for page in category_pages if (page - 1) in pages_needing_ocr)
        report["outputs"][category] = {
            "status": "written",
            "output_path": str(output_path),
            "resolved_ranges": [page_range.__dict__ for page_range in ranges],
            "ocr_risk_pages": ocr_risk_pages,
        }
        if tahsils is not None:
            report["outputs"][category]["tahsils"] = tahsils
        if ocr_risk_pages:
            report["warnings"].append(
                f"{category}: pages {', '.join(map(str, ocr_risk_pages))} may need OCR for text extraction."
            )

    return report


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Trim Uttar Pradesh 1971 PDFs into civic, medical, tehsil appendix, and village outputs."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--pdf-root", type=Path, default=DEFAULT_PDF_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--district", type=str, default=None)
    parser.add_argument("--sheet", choices=["Tehsil_and_Town", "Village_Level"], default=None)
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    (args.output / "reports").mkdir(parents=True, exist_ok=True)

    plan = load_trim_plan(args.workbook)
    districts = list(plan.values())
    if args.district:
        wanted = normalize_district(args.district)
        districts = [district for district in districts if normalize_district(district.district) == wanted]

    if not districts:
        raise SystemExit("No districts matched the current filters.")

    reports = []
    for district_plan in districts:
        print(f"Processing {district_plan.district}...")
        reports.append(process_district(district_plan, args.pdf_root, args.output, args.sheet))

    report_path = args.output / "reports" / "trim-report.json"
    report_path.write_text(json.dumps(reports, indent=2), encoding="utf-8")

    summary = []
    for report in reports:
        outputs = list(report["outputs"].values())
        summary.append(
            {
                "district": report["district"],
                "pdf_type": report["pdf_type"],
                "warnings": len(report["warnings"]),
                "written": sum(1 for output in outputs if output["status"] == "written"),
                "skipped": sum(1 for output in outputs if output["status"] != "written"),
            }
        )

    print(json.dumps(summary, indent=2))
    print(f"Report written to: {report_path}")


if __name__ == "__main__":
    main()
